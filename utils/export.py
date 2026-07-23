# -*- coding: utf-8 -*-
"""差旅搭子 - 导出工具：Excel 报销明细/汇总 + PDF 报销单。"""
from __future__ import annotations

from datetime import datetime
from typing import Optional, Dict, Any

from db.database import InvoiceDB
from db.models import Invoice
from ui.widgets import fmt_money, status_label

# ----------------------------- Excel ----------------------------- #
def export_excel(db: InvoiceDB, filepath: str, filters: Dict[str, Any] = None):
    """导出 Excel：明细表 + 汇总表。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    invoices = db.list_invoices(filters or {})
    wb = Workbook()

    # 明细表
    ws = wb.active
    ws.title = "报销明细"
    headers = ["序号", "发票号码", "分类", "开票日期", "销售方", "价税合计",
               "税额", "不含税金额", "所属行程", "状态", "来源", "备注"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="1976D2")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for i, inv in enumerate(invoices, 1):
        ws.append([
            i, inv.invoice_number, f"{inv.category_icon} {inv.category_name}",
            inv.issue_date, inv.vendor_name, round(inv.amount, 2),
            round(inv.tax_amount, 2), round(inv.amount_excl_tax, 2),
            inv.trip_name, status_label(inv.status), inv.source_label, inv.note,
        ])
    # 金额列数值格式
    for r in range(2, len(invoices) + 2):
        for col in (6, 7, 8):
            ws.cell(row=r, column=col).number_format = "#,##0.00"
    # 合计行
    total_row = len(invoices) + 2
    ws.cell(row=total_row, column=5, value="合计")
    ws.cell(row=total_row, column=6, value=round(sum(i.amount for i in invoices), 2))
    ws.cell(row=total_row, column=7, value=round(sum(i.tax_amount for i in invoices), 2))
    ws.cell(row=total_row, column=8, value=round(sum(i.amount_excl_tax for i in invoices), 2))
    for col in (6, 7, 8):
        ws.cell(row=total_row, column=col).number_format = "#,##0.00"
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    widths = [6, 22, 14, 12, 24, 12, 10, 12, 16, 10, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # 汇总表
    ws2 = wb.create_sheet("汇总")
    ws2.append(["差旅报销汇总"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append([])
    ws2.append(["按行程汇总"])
    ws2.append(["行程", "张数", "金额"])
    for r in ws2["A4:C4"]:
        for c in r:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head_fill
    for r in db.stats_by_trip():
        ws2.append([r["name"] or "(未归集)", r["cnt"], round(r["total"], 2)])
    ws2.append(["小计", len(invoices), round(sum(i.amount for i in invoices), 2)])
    ws2.append([])
    base = ws2.max_row + 1
    ws2.append(["按分类汇总"])
    ws2.append(["分类", "张数", "金额"])
    for r in ws2[f"A{base}:C{base}"]:
        for c in r:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head_fill
    for r in db.stats_by_category():
        if r["total"] > 0 or r["cnt"] > 0:
            ws2.append([f"{r['icon']} {r['name']}", r["cnt"], round(r["total"], 2)])
    for col in ("B", "C"):
        for row in range(1, ws2.max_row + 1):
            ws2[f"{col}{row}"].number_format = "#,##0.00" if col == "C" else "General"

    wb.save(filepath)
    return len(invoices)


# ----------------------------- PDF ----------------------------- #
def export_pdf(db: InvoiceDB, filepath: str, filters: Dict[str, Any] = None,
               title: str = "差旅报销单"):
    """导出 PDF 报销单（中文 CID 字体，无需额外字体文件）。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    FONT = "STSong-Light"

    invoices = db.list_invoices(filters or {})
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle("t", parent=styles["Title"], fontName=FONT, fontSize=18)
    style_norm = ParagraphStyle("n", parent=styles["Normal"], fontName=FONT, fontSize=9)
    style_small = ParagraphStyle("s", parent=styles["Normal"], fontName=FONT, fontSize=8)

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    elems = []
    elems.append(Paragraph(title, style_title))
    elems.append(Spacer(1, 4))
    elems.append(Paragraph(
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}　"
        f"发票张数：{len(invoices)}　"
        f"合计金额：{fmt_money(sum(i.amount for i in invoices))}", style_norm))
    elems.append(Spacer(1, 8))

    data = [["序号", "发票号码", "分类", "开票日期", "销售方", "价税合计", "税额", "行程"]]
    for i, inv in enumerate(invoices, 1):
        data.append([
            str(i), inv.invoice_number, inv.category_name, inv.issue_date,
            inv.vendor_name[:14], f"{inv.amount:.2f}", f"{inv.tax_amount:.2f}",
            inv.trip_name[:10],
        ])
    data.append(["", "", "", "", "合计",
                 f"{sum(i.amount for i in invoices):.2f}",
                 f"{sum(i.tax_amount for i in invoices):.2f}", ""])

    tbl = Table(data, repeatRows=1, colWidths=[
        12 * mm, 34 * mm, 24 * mm, 22 * mm, 40 * mm, 22 * mm, 18 * mm, 24 * mm])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ALIGN", (5, 0), (6, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef5ff")),
        ("FONTNAME", (0, -1), (-1, -1), FONT),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 10))
    elems.append(Paragraph("报销人签字：______________　　审核：______________　　日期：______________",
                           style_small))
    doc.build(elems)
    return len(invoices)
